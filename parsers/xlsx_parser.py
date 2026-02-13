"""Parser for XLSX files."""

from .base_parser import BaseParser


class XlsxParser(BaseParser):
    """Parser for XLSX files."""
    
    def parse(self, file_path: str) -> str:
        """
        Extract text from XLSX file.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Extracted text content
            
        Raises:
            Exception: If file cannot be parsed
        """
        try:
            import pandas as pd
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            text_parts = []
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                # Convert dataframe to string representation
                text_parts.append(f"Sheet: {sheet_name}\n{df.to_string()}")
            
            return '\n\n'.join(text_parts)
        except ImportError:
            try:
                # Fallback to openpyxl
                from openpyxl import load_workbook
                workbook = load_workbook(file_path, data_only=True)
                text_parts = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_text = []
                    for row in sheet.iter_rows(values_only=True):
                        row_text = ' '.join(str(cell) if cell is not None else '' for cell in row)
                        if row_text.strip():
                            sheet_text.append(row_text)
                    if sheet_text:
                        text_parts.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_text))
                
                return '\n\n'.join(text_parts)
            except ImportError:
                raise ImportError("pandas or openpyxl is required for XLSX parsing. Install with: pip install pandas openpyxl")
        except Exception as e:
            raise Exception(f"Error parsing XLSX file: {str(e)}")
